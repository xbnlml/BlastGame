#!/usr/bin/env python3
"""Contract for the production Excel import writer."""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import openpyxl

from tools.data.excel_writer import write_tiers


class ExcelWriterContractTest(unittest.TestCase):
    def _workbook(self, path: Path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["关卡", "难度", "档位", "胜率", "sd", "sc", "ratios", "of", "备注"])
        for index in range(5):
            sheet.append([58, "superhard" if index == 0 else "", f"Tier{index + 1}"])
        workbook.save(path)
        workbook.close()

    def test_normal_rows_preserve_shared_tier_configs(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "import.xlsx"
            self._workbook(path)
            tiers = [
                {"wr": 0.80, "sd": 1, "sc": 5, "ratios": "1,1,1,1,1", "of": 0.1, "note": "t1"},
                {"wr": 0.70, "sd": 2, "sc": 5, "ratios": "2,2,2,2,2", "of": 0.2, "note": "ignored"},
                {"wr": 0.60, "sd": 3, "sc": 5, "ratios": "3,3,3,3,3", "of": 0.3, "note": "t3"},
                {"wr": 0.45, "sd": 4, "sc": 5, "ratios": "4,4,4,4,4", "of": 0.4, "note": "t4"},
                {"wr": 0.30, "sd": 5, "sc": 5, "ratios": "5,5,5,5,5", "of": 0.5, "note": "ignored"},
            ]
            ok, message = write_tiers(
                58,
                tiers,
                targets={"diff": "normal"},
                xlsx_path=path,
            )
            self.assertTrue(ok, message)

            workbook = openpyxl.load_workbook(path, data_only=True)
            sheet = workbook.active
            rows = [[sheet.cell(row, column).value for column in range(4, 9)] for row in range(2, 7)]
            workbook.close()
            self.assertEqual(rows[0], rows[1])
            self.assertEqual(rows[3], rows[4])
            self.assertEqual(3, rows[2][1])
            self.assertEqual("3,3,3,3,3", rows[2][3])

    def test_reimport_does_not_load_production_code_from_archive(self):
        source = (Path(__file__).resolve().parents[2] / "tools" / "reimport.py").read_text(encoding="utf-8")
        self.assertNotIn("project-state/_archive/write_excel.py", source)
        self.assertNotIn("_import_write_excel", source)
        self.assertIn("from tools.data.excel_writer import write_tiers", source)

    def test_reimport_db_sync_uses_write_tool_payload_contract(self):
        source = (Path(__file__).resolve().parents[2] / "tools" / "reimport.py").read_text(encoding="utf-8")
        self.assertIn("'_write_payload.json'", source)
        self.assertNotIn("'_reimport_payload.json'", source)


if __name__ == "__main__":
    unittest.main(verbosity=2)
