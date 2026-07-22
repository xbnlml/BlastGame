"""
BlastGame Excel import template.
Copy this and fill in the level-specific values.

Excel format:
  A: 关号 (合并5行)  |  B: 难度 (合并5行)  |  C: 档位  |  D: 通关率  |  E: SD  |  F: SC  |  G: 参数  |  H: OF
"""

import openpyxl

REPO = r"C:\Users\Administrator\Documents\BlastGame"
XL_PATH = REPO + "/Doc/手动挑配置记录.xlsx"

LEVEL = 80      # <-- change this
DIFFICULTY = "hard"  # <-- "normal" or "hard"

# 5 tiers: (tier_name, winrate_decimal, sd, sc, ratios, of)
TIERS = [
    ("Tier1", 0.808, 39, 5, "10,10,10,1,1", 0.5),
    ("Tier2", 0.59,  34, 5, "10,10,1,1,1", 0.5),
    ("Tier3", 0.392, 10, 1, "10",          0.107),
    ("Tier4", 0.25,  20, 5, "1,1,1,10,1",  0.5),
    ("Tier5", 0.14,  20, 5, "1,1,1,1,10",  0.5),
]

wb = openpyxl.load_workbook(XL_PATH)
ws = wb.active

for row_idx in range(2, ws.max_row + 1):
    v = ws.cell(row=row_idx, column=1).value
    if v and str(v).strip() == str(LEVEL):
        for i, (tier, wr, sd, sc, ratios, ofv) in enumerate(TIERS):
            r = row_idx + i
            ws.cell(row=r, column=2).value = DIFFICULTY  # only writes on first row (merged)
            ws.cell(row=r, column=3).value = tier
            ws.cell(row=r, column=4).value = wr
            ws.cell(row=r, column=5).value = sd
            ws.cell(row=r, column=6).value = sc
            ws.cell(row=r, column=7).value = ratios
            ws.cell(row=r, column=8).value = ofv
        break

wb.save(XL_PATH)
print(f"L{LEVEL} imported")
