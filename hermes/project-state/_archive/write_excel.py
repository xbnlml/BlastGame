#!/usr/bin/env python3
"""写入 5 档配置到 Excel，自动处理 Normal 关 T2=T1 / T5=T4。
写入后立即验证 T1-T5 非空，防止错位。
"""
import os, sys

TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
XL_PATH = os.path.join(TOOL_DIR, '..', '手动挑配置记录.xlsx')


def write_tiers(lv, tiers, targets=None):
    """向 Excel 写入 5 档配置。tiers = [dict, dict, ...] 5个。
    Normal 关自动展开 T2=T1, T5=T4。
    """
    from tools.data.adapters import excel_target as et
    import openpyxl

    # 展开为完整 5 行
    if targets is None:
        t = et.get_target(lv)
        targets = t

    rows = []
    diff = targets['diff'] if targets else 'normal'

    for i, x in enumerate(tiers):
        if diff == 'normal':
            if i == 1:  # T2 = T1
                rows.append(dict(rows[0]))
                rows[1]['note'] = ''
                continue
            if i == 4:  # T5 = T4
                rows.append(dict(rows[3]))
                rows[4]['note'] = ''
                continue
        rows.append(dict(x))

    # 补齐到 5 行
    while len(rows) < 5:
        rows.append(dict(rows[-1]))

    wb = openpyxl.load_workbook(XL_PATH)
    ws = wb.active

    # 找到该关卡的起始行
    start_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == lv:
            start_row = row
            break
    if start_row is None:
        wb.close()
        return False, 'Excel 中找不到 L{}'.format(lv)

    # 写入 5 行，每行列全部填充
    for i, d in enumerate(rows):
        r = start_row + i
        ws.cell(r, 1, lv)
        ws.cell(r, 4, round(d.get('wr', 0), 3))
        ws.cell(r, 5, int(d.get('sd', 0)))
        ws.cell(r, 6, int(d.get('sc', 5)))
        ws.cell(r, 7, str(d.get('ratios', '')))
        ws.cell(r, 8, float(d.get('of', 0.5)))
        ws.cell(r, 9, d.get('note', ''))

    wb.save(XL_PATH)
    wb.close()

    # 写入后验证：5 行全部非空
    wb2 = openpyxl.load_workbook(XL_PATH)
    ws2 = wb2.active
    errors = []
    for i in range(5):
        r = start_row + i
        lv_v = ws2.cell(r, 1).value
        sd_v = ws2.cell(r, 5).value
        ratios_v = ws2.cell(r, 7).value
        if lv_v is None or sd_v is None or ratios_v is None:
            errors.append('T{} 为空 (row {})'.format(i + 1, r))
    wb2.close()
    if errors:
        return False, '; '.join(errors)

    return True, 'OK ({} tiers, T1 sd={}, T5 sd={})'.format(
        len(rows), rows[0]['sd'], rows[4]['sd'])
