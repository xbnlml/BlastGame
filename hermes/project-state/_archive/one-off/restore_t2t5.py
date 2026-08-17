"""从 _append_block_backup.json 恢复 15 关 T2-T5 到原位行（T1 已写，跳过）。"""
import openpyxl, json

XL = r'D:\download\BlastGame\hermes\手动挑配置记录.xlsx'
data = json.load(open(r'D:\download\BlastGame\hermes\_append_block_backup.json', encoding='utf-8'))

wb = openpyxl.load_workbook(XL)
ws = wb.active
print(f'总行数: {ws.max_row}')

def find_start(ws, lv):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == lv:
            return row
    return None

fixed = 0
for lv_s, tiers in sorted(data.items(), key=lambda x: int(x[0])):
    lv = int(lv_s)
    start = find_start(ws, lv)
    if start is None:
        print(f'  L{lv}: 原位找不到!')
        continue
    for i, t in enumerate(tiers):
        r = start + i
        # 只补 T2-T5（T1 已写回）；档位列统一 TierN
        tier = str(t['tier'])
        if not tier.startswith('Tier'):
            tier = 'Tier' + tier
        ws.cell(r, 3, tier)
        ws.cell(r, 4, t['wr'])
        ws.cell(r, 5, t['sd'])
        ws.cell(r, 6, t['sc'])
        ws.cell(r, 7, t['ratios'])
        ws.cell(r, 8, t['of'])
        ws.cell(r, 9, t['note'])
        fixed += 1
    print(f'  L{lv}: R{start}-{start+4} 补全 {len(tiers)} 档')

wb.save(XL)
print(f'\n保存完成, 补全 {fixed} 行, 总行数: {ws.max_row}')
wb.close()
