"""修复 4 关（156/166/183/200）胜率格式：百分数 → 小数，并对照 summary 真源。"""
import openpyxl, csv, os

XL = r'D:\download\BlastGame\hermes\手动挑配置记录.xlsx'
batch = r'C:\Users\Administrator\Documents\BlastGame\telemetry\multi-tier-opt\152_156_166-167_183_189_200-2026-08-04T10-49-51'

def norm(r):
    return str(r or '').strip().replace('，', ',')

# summary 真源（胜率原始值，summary.csv 是小数）
truth = {}
for lv in [156, 166, 183, 200]:
    d = [x for x in os.listdir(batch) if x.startswith(f'{lv}-')][0]
    with open(os.path.join(batch, d, 'summary.csv'), encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    truth[lv] = []
    for r in rows:
        truth[lv].append({
            'tier': 'Tier' + r['Tier'].split('-')[0][-1],
            'wr_raw': float(r['VerifiedWinRate']),      # 小数
            'wr_pct': round(float(r['VerifiedWinRate']) * 100, 2),  # 百分数
            'sd': int(float(r['StartDifficulty'])),
            'sc': int(float(r['ShuffleSplitCount'])),
            'ratios': norm(r['ShuffleSplitRatios']),
            'of': float(r['ShuffleOverflowFactor']),
            'games': r['TotalRuns'],
        })
    print(f'L{lv} summary 真源:')
    for t in truth[lv]:
        wr_pct = t['wr_pct']
        print(f'  {t["tier"]}: WR={t["wr_raw"]:.4f} (={wr_pct}%) sd={t["sd"]} sc={t["sc"]} ratios={t["ratios"]} of={t["of"]} {t["games"]}局')

wb = openpyxl.load_workbook(XL)
ws = wb.active

def find_start(lv):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == lv:
            return row
    return None

print()
for lv in [156, 166, 183, 200]:
    start = find_start(lv)
    assert start is not None, f'L{lv} 原位缺失'
    print(f'=== L{lv} 修复 ===')
    for i, t in enumerate(truth[lv]):
        r = start + i
        cur = ws.cell(r, 4).value
        new = t['wr_raw']  # 小数
        ws.cell(r, 4, new)
        # 校验其他列
        assert ws.cell(r, 3).value == t['tier'], f'L{lv} 档位: {ws.cell(r,3).value} vs {t["tier"]}'
        assert ws.cell(r, 5).value == t['sd'], f'L{lv} sd: {ws.cell(r,5).value} vs {t["sd"]}'
        assert ws.cell(r, 6).value == t['sc'], f'L{lv} sc: {ws.cell(r,6).value} vs {t["sc"]}'
        assert str(ws.cell(r, 7).value).replace(' ', '') == t['ratios'].replace(' ', ''), f'L{lv} ratios: {ws.cell(r,7).value} vs {t["ratios"]}'
        assert abs(float(ws.cell(r, 8).value) - t['of']) < 1e-6, f'L{lv} of: {ws.cell(r,8).value} vs {t["of"]}'
        print(f'  {t["tier"]}: 胜率 {cur} → {new} (其他列已核对一致)')

wb.save(XL)
print(f'\n保存完成')
wb.close()
