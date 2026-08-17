"""修复 Excel 入库记录格式：
1. 删除误追加的 75 行（R1002-1076，15 关）
2. 15 关 + 4 关全部写回原位行（保留原格式，覆盖值+备注）
3. 档位列统一 'Tier1'-'Tier5' 格式
"""
import openpyxl, csv, os, shutil
from datetime import datetime

XL = r'D:\download\BlastGame\hermes\手动挑配置记录.xlsx'
batch = r'C:\Users\Administrator\Documents\BlastGame\telemetry\multi-tier-opt\152_156_166-167_183_189_200-2026-08-04T10-49-51'

# 备份
bak = XL.replace('.xlsx', f'_before_format_fix_{datetime.now().strftime("%H%M%S")}.bak')
shutil.copy2(XL, bak)
print(f'备份: {bak}')

def norm(r):
    return str(r or '').strip().replace('，', ',')

# ============ 收集所有要写的数据 ============
# 15 关（8-04 之前入库，来自池子/asset 匹配）—— 从追加块读取原值
# 4 关（8-04 新 summary）—— 从 summary.csv 读取
new_data = {}  # lv -> {diff, [(tier, wr, sd, sc, ratios, of, note), ...]}

# 4 关 summary
diff_map = {156: 'Normal', 166: 'Hard', 183: 'Hard', 200: 'SuperHard'}
for lv in [156, 166, 183, 200]:
    d = [x for x in os.listdir(batch) if x.startswith(f'{lv}-')][0]
    with open(os.path.join(batch, d, 'summary.csv'), encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    tiers = []
    for r in rows:
        tiers.append(('Tier' + r['Tier'].split('-')[0][-1],  # Tier1
                      round(float(r['VerifiedWinRate']) * 100, 2),
                      int(float(r['StartDifficulty'])),
                      int(float(r['ShuffleSplitCount'])),
                      norm(r['ShuffleSplitRatios']),
                      float(r['ShuffleOverflowFactor']),
                      f'多档位summary 2026-08-04 {r["TotalRuns"]}局'))
    new_data[lv] = (diff_map[lv], tiers)

# 15 关（153/155/158/159/163/168/172/174/175/184/186/187/194/197/199）
# 从当前 Excel 的追加块读取（R1002-1076），然后删除追加块并写回原位
wb = openpyxl.load_workbook(XL)
ws = wb.active

# 先读追加块数据
append_data = {}  # lv -> [(tier, wr, sd, sc, ratios, of, note)]
for r in range(1002, ws.max_row + 1):
    lv = ws.cell(r, 1).value
    if lv is None:
        continue
    lv = int(lv)
    if lv not in append_data:
        append_data[lv] = []
    note = ws.cell(r, 9).value
    append_data[lv].append((
        'Tier' + str(ws.cell(r, 3).value),  # 数字→TierN
        ws.cell(r, 4).value,   # wr (小数)
        ws.cell(r, 5).value,   # sd
        ws.cell(r, 6).value,   # sc
        ws.cell(r, 7).value,   # ratios
        ws.cell(r, 8).value,   # of
        note,
    ))
print(f'追加块读取: {len(append_data)} 关')
for lv, tiers in sorted(append_data.items()):
    print(f'  L{lv}: {len(tiers)} 档, 备注={tiers[0][6][:30] if tiers[0][6] else None}')

# ============ 先写回原位（数据还在追加块里，可随时读） ============
def find_start(ws, lv):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == lv:
            return row
    return None

# 15 关写回（难度列从原位保留，值更新）
for lv, tiers in sorted(append_data.items(), key=lambda x: int(x[0])):
    start = find_start(ws, lv)
    if start is None:
        print(f'  L{lv}: 原位找不到!')
        continue
    for i, (tier, wr, sd, sc, ratios, of, note) in enumerate(tiers):
        r = start + i
        ws.cell(r, 3, tier)
        ws.cell(r, 4, wr)
        ws.cell(r, 5, sd)
        ws.cell(r, 6, sc)
        ws.cell(r, 7, ratios)
        ws.cell(r, 8, of)
        ws.cell(r, 9, note)
    print(f'  L{lv}: 写回原位 R{start}-{start+4}')

# 4 关写回（含难度列）
for lv in [156, 166, 183, 200]:
    diff, tiers = new_data[lv]
    start = find_start(ws, lv)
    if start is None:
        print(f'  L{lv}: 原位找不到!')
        continue
    for i, (tier, wr, sd, sc, ratios, of, note) in enumerate(tiers):
        r = start + i
        if i == 0:
            ws.cell(r, 2, diff)
        ws.cell(r, 3, tier)
        ws.cell(r, 4, wr)
        ws.cell(r, 5, sd)
        ws.cell(r, 6, sc)
        ws.cell(r, 7, ratios)
        ws.cell(r, 8, of)
        ws.cell(r, 9, note)
    print(f'  L{lv}: 写回原位 R{start}-{start+4}')

# ============ 确认写回成功后，再删除追加块 ============
# 写回验证：原位块数据与追加块一致
verify_ok = True
for lv, tiers in sorted(append_data.items(), key=lambda x: int(x[0])):
    start = find_start(ws, lv)
    if start is None:
        verify_ok = False
        continue
    for i, (tier, wr, sd, sc, ratios, of, note) in enumerate(tiers):
        r = start + i
        if ws.cell(r, 3).value != tier or ws.cell(r, 4).value != wr:
            verify_ok = False
            print(f'  L{lv} T{i+1}: 写回验证失败!')
if not verify_ok:
    print('!! 写回验证失败，中止删除追加块（数据仍在追加块中）')
    wb.save(XL)
    wb.close()
    raise SystemExit(1)
print('写回验证通过，现在删除追加块')

for _ in range(ws.max_row - 1001):
    ws.delete_rows(1002)
print(f'删除追加块后总行数: {ws.max_row}')

wb.save(XL)
print(f'\n保存完成, 总行数: {ws.max_row}')
wb.close()
