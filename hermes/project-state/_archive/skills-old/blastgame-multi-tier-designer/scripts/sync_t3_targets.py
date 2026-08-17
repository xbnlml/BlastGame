#!/usr/bin/env python3
"""
将 lv_win_config_test.xlsx 的 T3 目标胜率同步到 1-200关设定.xlsx 的 D 列。
用法：python sync_t3_targets.py
"""
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parents[4]
DST = REPO / "Doc/1-200关设定.xlsx"
SRC = REPO / "Assets/LvEditorConfig/lv_win_config_test.xlsx"

# Read T3 targets
src = openpyxl.load_workbook(str(SRC), data_only=True)
sws = src.active
t3_targets = {}
for row in sws.iter_rows(min_row=5, values_only=True):
    lv = row[0]
    t3 = row[3]
    if lv is not None and isinstance(lv, (int, float)):
        t3_targets[int(lv)] = t3
src.close()

# Update 1-200关设定
wb = openpyxl.load_workbook(str(DST))
ws = wb['1-200关设定']
updated = 0
for r in range(2, ws.max_row + 1):
    b = ws.cell(r, 2).value
    if b is not None:
        try:
            lv = int(b)
            if lv in t3_targets:
                ws.cell(r, 4).value = t3_targets[lv]
                updated += 1
        except ValueError:
            pass
wb.save(str(DST))
wb.close()
print(f"Synced {updated} T3 targets to {DST.name}")
