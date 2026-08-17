#!/usr/bin/env python3
"""
将 1-200关设定.xlsx 的难度列同步到其他三个文件（lv_win_config_test, asset, 手动挑配置记录）。
用法：python sync_difficulty_across_files.py [--dry-run]
默认 dry-run，加 --apply 才实际写入。
"""
import argparse
import re
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parents[4]
STD = REPO / "Doc/1-200关设定.xlsx"
LV_WIN = REPO / "Assets/LvEditorConfig/lv_win_config_test.xlsx"
RECORD = REPO / "Doc/手动挑配置记录.xlsx"
ASSET_DIR = REPO / "Assets/GameModule/GameMain/ConfigSo/Generated_enum/test"
DL_MAP = {"normal": 0, "hard": 1, "superhard": 2}
DL_MAP_REV = {0: "normal", 1: "hard", 2: "superhard"}


def read_std() -> dict[int, str]:
    """Read authoritative difficulties from 1-200关设定."""
    wb = openpyxl.load_workbook(str(STD), data_only=True)
    ws = wb["1-200关设定"]
    out = {}
    for r in range(2, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b is not None:
            try:
                out[int(b)] = ws.cell(r, 3).value
            except (ValueError, TypeError):
                pass
    wb.close()
    return out


def fix_lv_win_config(std: dict[int, str], apply: bool) -> int:
    """Fix difficulty column I in lv_win_config_test.xlsx."""
    wb = openpyxl.load_workbook(str(LV_WIN))
    ws = wb.active
    fixed = 0
    for row in ws.iter_rows(min_row=5):
        lv_cell = row[0]
        diff_cell = row[8]
        if lv_cell.value is not None and isinstance(lv_cell.value, (int, float)):
            lv = int(lv_cell.value)
            expected = std.get(lv)
            if expected and diff_cell.value != expected:
                if apply:
                    diff_cell.value = expected
                fixed += 1
    if apply:
        wb.save(str(LV_WIN))
    wb.close()
    return fixed


def fix_assets(std: dict[int, str], apply: bool) -> int:
    """Fix difficultyLevel in test/*.asset files."""
    fixed = 0
    for lv, expected_diff in std.items():
        path = ASSET_DIR / f"{lv}.asset"
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8")
        m = re.search(r"difficultyLevel:\s*(\d+)", text)
        if not m:
            continue
        current_dl = int(m.group(1))
        current_diff = DL_MAP_REV[current_dl]
        expected_dl = DL_MAP.get(expected_diff)
        if expected_dl is not None and current_dl != expected_dl:
            if apply:
                new_text = text.replace(
                    f"difficultyLevel: {current_dl}",
                    f"difficultyLevel: {expected_dl}",
                )
                path.write_text(new_text, encoding="utf-8")
            fixed += 1
    return fixed


def fix_record(std: dict[int, str], apply: bool) -> int:
    """Fix difficulty column B in 手动挑配置记录.xlsx."""
    wb = openpyxl.load_workbook(str(RECORD))
    ws = wb.active
    fixed = 0
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is not None:
            try:
                lv = int(str(a).strip())
                expected = std.get(lv)
                if expected:
                    current = ws.cell(r, 2).value
                    if current and str(current).lower().strip() != expected:
                        if apply:
                            ws.cell(r, 2).value = expected
                        fixed += 1
            except ValueError:
                pass
    if apply:
        wb.save(str(RECORD))
    wb.close()
    return fixed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    std = read_std()
    print(f"1-200关设定: {len(std)} levels")

    n1 = fix_lv_win_config(std, args.apply)
    n2 = fix_assets(std, args.apply)
    n3 = fix_record(std, args.apply)

    print(f"lv_win_config_test.xlsx: {n1} mismatches")
    print(f"Asset files:             {n2} mismatches")
    print(f"手动挑配置记录.xlsx:        {n3} mismatches")
    print(f"Total: {n1 + n2 + n3}")
    if not args.apply:
        print("DRY-RUN: pass --apply to write changes")
    else:
        print("Changes applied.")


if __name__ == "__main__":
    main()
