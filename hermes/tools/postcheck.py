#!/usr/bin/env python3
"""操作后验证 — 入库/改关卡/提交后的自检。

用法:
  python tools/postcheck.py 入库 62
  python tools/postcheck.py 改关卡 59
  python tools/postcheck.py 提交 -l 59,81,82 --tag l6-r2
"""
import os, sys, json, subprocess, glob, re, csv

TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
STAGE_DIR = os.path.join(TOOL_DIR, '..', 'stage-data')
BOARD_FILE = os.path.join(TOOL_DIR, '..', 'project-state', 'board.md')
REPO = os.environ.get('BLASTGAME_REPO', r'C:\Users\Administrator\Documents\BlastGame')

def get_asset_path(lv):
    for root, dirs, files in os.walk(os.path.join(REPO, 'Assets/GameModule/GameMain/ConfigSo/Generated_enum/test')):
        if f'{lv}.asset' in files:
            return os.path.join(root, f'{lv}.asset')
    return os.path.join(REPO, 'Assets/GameModule/GameMain/ConfigSo/Generated_enum/test', f'{lv}.asset')

def get_excel_tiers(lv):
    import openpyxl
    xl = os.path.join(TOOL_DIR, '..', '手动挑配置记录.xlsx')
    if not os.path.exists(xl): return None
    wb = openpyxl.load_workbook(xl)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == lv:
            tiers = []
            for i in range(5):
                r = row + i
                ratios = ws.cell(r, 7).value
                if ratios is None:
                    tiers.append(tiers[0] if i == 1 else tiers[3])
                else:
                    tiers.append({
                        'sd': int(ws.cell(r, 5).value) if ws.cell(r, 5).value else 0,
                        'sc': int(ws.cell(r, 6).value) if ws.cell(r, 6).value else 5,
                        'ratios': str(ratios).strip(),
                        'of': float(ws.cell(r, 8).value) if ws.cell(r, 8).value else 0.5
                    })
            return tiers
    return None


if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='操作后验证')
    sp = ap.add_subparsers(dest='cmd', required=True)
    p1 = sp.add_parser('入库')
    p1.add_argument('lv', type=int)
    p2 = sp.add_parser('改关卡')
    p2.add_argument('lv', type=int)
    p3 = sp.add_parser('提交')
    p3.add_argument('-l', '--levels', required=True)
    p3.add_argument('--tag', default='')

    args = ap.parse_args()
    has_error = False

    if args.cmd == '入库':
        lv = args.lv
        sys.path.insert(0, os.path.join(TOOL_DIR, 'data', 'adapters'))

        # 1. Excel T1-T5 非空
        ex = get_excel_tiers(lv)
        if not ex:
            print('❌ Excel 中找不到 L{}'.format(lv))
            has_error = True
        elif len(ex) != 5:
            print('❌ Excel L{} 仅 {} 行（应 5）'.format(lv, len(ex)))
            has_error = True
        else:
            print('✅ Excel L{} 5 行完整'.format(lv))

            # 2. Normal 关 T1=T2, T4=T5
            from tools.data.adapters import excel_target as et
            t = et.get_target(lv)
            if t and t['diff'] == 'normal':
                t1, t2 = int(ex[0]['sd']), int(ex[1]['sd'])
                t4, t5 = int(ex[3]['sd']), int(ex[4]['sd'])
                if t1 != t2:
                    print('❌ Normal 关 T1(sd={}) ≠ T2(sd={})'.format(t1, t2))
                    has_error = True
                if t4 != t5:
                    print('❌ Normal 关 T4(sd={}) ≠ T5(sd={})'.format(t4, t5))
                    has_error = True
                if t1 == t2 and t4 == t5:
                    print('✅ Normal 结构正确')

        # 3. Asset vs Excel 一致性
        from tools.asset_patcher import read_ddc as rd
        cur = rd(lv)
        if not cur:
            print('❌ asset 读不出')
            has_error = True
        elif ex and len(ex) == 5:
            ok = True
            for i in range(5):
                if int(ex[i]['sd']) != int(cur[i]['sd']) or str(ex[i]['ratios']) != str(cur[i]['ratios']):
                    print('❌ L{} T{} Excel(sd={} r={}) ≠ Asset(sd={} r={})'.format(
                        lv, i+1, ex[i]['sd'], ex[i]['ratios'], cur[i]['sd'], cur[i]['ratios']))
                    ok = False
                    has_error = True
            if ok:
                print('✅ Asset vs Excel 一致')

        # 4. Snapshot 备份存在
        snap_dir = os.path.join(TOOL_DIR, '..', 'asset_backups')
        snap_files = glob.glob(os.path.join(snap_dir, 'snapshot-*', '{}.asset'.format(lv)))
        if snap_files:
            print('✅ Snapshot 备份: {}'.format(os.path.basename(os.path.dirname(snap_files[-1]))))
        else:
            print('⚠️ 无 snapshot 备份')
            has_error = True

        # 5. Board 总数
        if os.path.exists(BOARD_FILE):
            with open(BOARD_FILE) as f:
                content = f.read()
            total = 0
            import re as re2
            for m in re2.finditer(r'(\d+[\d,]*)', content):
                nums = [int(x) for x in re2.findall(r'\d+', m.group(0))]
                for n in nums:
                    if 51 <= n <= 200:
                        total += 1
            if total == 150:
                print('✅ Board 总数 150')
            else:
                print('⚠️ Board 仅 {} 关（应 150）'.format(total))

    elif args.cmd == '改关卡':
        lv = args.lv
        slv = str(lv)

        # 检查 retired_levels.json
        retired_file = os.path.join(TOOL_DIR, '..', 'project-state', 'retired_levels.json')
        if os.path.exists(retired_file):
            with open(retired_file) as f:
                rdata = json.load(f)
            if slv in rdata:
                print('✅ retired_levels.json 已记录')
            else:
                print('❌ retired_levels.json 未记录')
                has_error = True

        # 检查 _last_refresh.json 时间防线
        tracking = os.path.join(STAGE_DIR, '_last_refresh.json')
        if os.path.exists(tracking):
            with open(tracking, encoding='utf-8') as f:
                tdata = json.load(f)
            if slv in tdata.get('asset_updated_at', {}):
                print('✅ 时间防线已设: {}'.format(tdata['asset_updated_at'][slv]))
            else:
                print('❌ 时间防线未设')
                has_error = True

        # 检查 stage-data
        sd_file = os.path.join(STAGE_DIR, slv, '{}.bot.json'.format(slv))
        if os.path.exists(sd_file):
            with open(sd_file) as f:
                bot_data = json.load(f)
            if len(bot_data) == 0:
                print('✅ stage-data 已清空')
            else:
                print('⚠️ stage-data 仍有 {} 条记录（时间防线过滤后）'.format(len(bot_data)))

        if has_error:
            print('❌ 验证未通过')
        else:
            print('✅ 全部通过')

    elif args.cmd == '提交':
        errors = []
        # 检查批跑 bot 目录
        bot_dir = os.path.join(REPO, 'telemetry', 'bot')
        if os.path.isdir(bot_dir):
            dirs = sorted([d for d in os.listdir(bot_dir) if not d.startswith('_')])
            latest = dirs[-1] if dirs else ''
            if latest:
                dp = os.path.join(bot_dir, latest)
                tiers_ok = sum(1 for td in os.listdir(dp) if os.path.isdir(os.path.join(dp, td)) and re.search(r'T\d+', td))
                if tiers_ok >= 5:
                    print('✅ 最新批跑 {} 有 {} 档'.format(latest, tiers_ok))
                else:
                    print('⚠️ 最新批跑 {} 仅 {} 档'.format(latest, tiers_ok))
        for e in errors:
            print('❌ ' + e)
        if errors:
            sys.exit(1)
