import sys; sys.path.insert(0, '.')
from tools.data import pool
import openpyxl

wb = openpyxl.load_workbook('C:/Users/Administrator/Documents/BlastGame/Assets/LvEditorConfig/lv_win_config_test.xlsx', data_only=True)
ws = wb.active
targets = {}
for r in range(2, ws.max_row + 1):
    lv = str(ws.cell(r, 1).value)
    if lv in ['172','173','174','175','176','178','179','180','181','184']:
        targets[lv] = [float(ws.cell(r,c).value or 0)*100 for c in range(2,7)]

levels = ['172','173','174','175','176','178','179','180','181','184']

def norm_of(v):
    try: return f'{float(v):.3f}'.rstrip('0').rstrip('.')
    except: return str(v).strip()

print('{:4} {:10} {:>7} {:>7} {:>7} {:>7} {:>7}  gaps'.format('关','难度','T1','T2','T3','T4','T5'))
print('-' * 60)

for lv in levels:
    recs = pool.dedup_records(pool.get_preferred_records(lv))
    tgt = targets.get(lv, [0]*5)
    is_normal = (tgt[0] == tgt[1])
    
    # Dedup by config, only bot/summary/phase2 (games>=100)
    uniq = {}
    for r in recs:
        wr = r.get('wr', 0)
        sd = r.get('sd')
        sc = r.get('sc')
        ratios = str(r.get('ratios','')).replace(' ','')
        of_v = norm_of(str(r.get('of','')))
        src = r.get('source','?')
        games = r.get('totalGames',0)
        if games < 100: continue
        key = f'{sd}|{sc}|{ratios}|{of_v}'
        prio = {'bot':0,'summary':1,'phase2':5,'phase1':10}.get(src, 10)
        if key not in uniq or prio < uniq[key]['prio']:
            uniq[key] = {'wr':wr, 'sd':sd, 'sc':sc, 'ratios':ratios, 'of':of_v, 'prio':prio}
    
    items = sorted(uniq.values(), key=lambda x: -x['wr'])
    n = len(items)
    
    if is_normal:
        best, best_score = None, None
        for i in range(n):
            for j in range(n):
                w2 = items[j]['wr']
                if w2 > items[i]['wr']: continue
                for k in range(n):
                    w3 = items[k]['wr']
                    if w3 > w2: continue
                    if i==j==k: continue
                    s = abs(items[i]['wr']-tgt[0])*2 + abs(w2-tgt[2]) + abs(w3-tgt[3])*2
                    s += (items[i]['prio']+items[j]['prio']+items[k]['prio'])*0.1
                    if best is None or s < best_score:
                        best = (items[i], items[j], items[k])
                        best_score = s
        if best:
            wrs = [best[0]['wr']]*2 + [best[1]['wr']] + [best[2]['wr']]*2
        else:
            wrs = [None]*5
    else:
        best, best_score = None, None
        for a in range(n):
            for b in range(a+1, n):
                for c in range(b+1, n):
                    for d in range(c+1, n):
                        for e in range(d+1, n):
                            w = [items[x]['wr'] for x in [a,b,c,d,e]]
                            s = sum(abs(w[i]-tgt[i]) for i in range(5))
                            s += sum(items[x]['prio'] for x in [a,b,c,d,e])*0.1
                            if best is None or s < best_score:
                                best = [items[x] for x in [a,b,c,d,e]]
                                best_score = s
        wrs = [c['wr'] for c in best] if best else [None]*5
    
    ws_str = ' '.join('{:>6.1f}%'.format(w) if w else '  N/A  ' for w in wrs)
    gaps = []
    for i in range(4):
        if wrs[i] and wrs[i+1]:
            gaps.append('{:.1f}'.format(wrs[i]-wrs[i+1]))
        else:
            gaps.append('N/A')
    
    print('L{:<4} {:<10} {}  {}'.format(lv, 'Normal' if is_normal else ('Hard' if tgt[0]>=70 else 'S.Hard'), ws_str, '/'.join(gaps)))
