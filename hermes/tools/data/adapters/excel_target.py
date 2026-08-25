"""lv_win_config_test.xlsx 目标胜率读取

统一入口，自动处理格式变体。
"""
import os
from openpyxl import load_workbook

REPO = os.environ.get('BLASTGAME_REPO', os.path.join(os.path.expanduser('~'), 'Documents', 'BlastGame'))
XLSX = os.path.join(REPO, 'Assets/LvEditorConfig/lv_win_config_test.xlsx')

_CACHE = None  # 2026-08-06: 模块级缓存——150 关循环每关 load_workbook 是性能瓶颈（~135s）


def read_targets(levels=None):
    """读目标胜率表，返回 {lv: {'diff': 'normal', 'tiers': [T1%..T5%]}}"""
    global _CACHE
    if _CACHE is None:
        wb = load_workbook(XLSX, data_only=True)
        ws = wb.active
        targets = {}
        prev_lv = None
        for r in range(3, ws.max_row + 1):
            lv_raw = ws.cell(r, 1).value
            if lv_raw is not None:
                prev_lv = int(lv_raw)
            if prev_lv is None:
                continue
            t1 = ws.cell(r, 2).value
            if t1 is None:
                continue
            if prev_lv not in targets:
                tiers = []
                for c in range(2, 7):
                    v = ws.cell(r, c).value
                    tiers.append(float(v) * 100 if v else 0)
                targets[prev_lv] = {
                    'diff': str(ws.cell(r, 9).value or '').lower(),
                    'tiers': tiers,
                }
        wb.close()
        _CACHE = targets

    if levels is not None:
        return {k: v for k, v in _CACHE.items() if k in levels}
    return _CACHE


def get_target(lv):
    """快捷方式"""
    t = read_targets({int(lv)})
    return t.get(int(lv), {})
