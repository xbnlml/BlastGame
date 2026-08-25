"""Write finalized tier configurations to the import-record workbook."""
from __future__ import annotations

from pathlib import Path
from typing import Any


HERMES = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = HERMES / "手动挑配置记录.xlsx"


def write_tiers(
    level: int,
    tiers: list[dict[str, Any]],
    targets: dict[str, Any] | None = None,
    *,
    xlsx_path: str | Path | None = None,
) -> tuple[bool, str]:
    """Update five existing workbook rows and verify required cells after save.

    Normal levels share T1/T2 and T4/T5 physical configurations. The caller
    still supplies five tier slots; this function enforces the sharing rule at
    the final workbook seam.
    """
    from openpyxl import load_workbook
    from tools.data.adapters import excel_target

    target_info = targets if targets is not None else excel_target.get_target(level)
    difficulty = target_info["diff"] if target_info else "normal"
    rows: list[dict[str, Any]] = []
    for index, tier in enumerate(tiers):
        if difficulty == "normal" and index == 1:
            if not rows:
                return False, "Normal T2 cannot be expanded before T1"
            rows.append(dict(rows[0]))
            rows[-1]["note"] = ""
            continue
        if difficulty == "normal" and index == 4:
            if len(rows) < 4:
                return False, "Normal T5 cannot be expanded before T4"
            rows.append(dict(rows[3]))
            rows[-1]["note"] = ""
            continue
        rows.append(dict(tier))
    if len(rows) != 5:
        return False, f"expected five tier rows, got {len(rows)}"

    path = Path(xlsx_path) if xlsx_path is not None else DEFAULT_XLSX
    try:
        workbook = load_workbook(path)
    except Exception as exc:
        return False, f"cannot open workbook {path}: {exc}"
    sheet = workbook.active
    start_row = next(
        (row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value == level),
        None,
    )
    if start_row is None:
        workbook.close()
        return False, f"Excel 中找不到 L{level}"

    for index, data in enumerate(rows):
        row = start_row + index
        sheet.cell(row, 1, level)
        sheet.cell(row, 4, round(float(data.get("wr", 0)), 4))
        sheet.cell(row, 5, int(data.get("sd", 0)))
        sheet.cell(row, 6, int(data.get("sc", 5)))
        sheet.cell(row, 7, str(data.get("ratios", "")))
        sheet.cell(row, 8, float(data.get("of", 0.5)))
        sheet.cell(row, 9, str(data.get("note", "")))
    try:
        workbook.save(path)
    except Exception as exc:
        workbook.close()
        return False, f"cannot save workbook {path}: {exc}"
    workbook.close()

    verify = load_workbook(path, data_only=True)
    verify_sheet = verify.active
    errors = []
    for index in range(5):
        row = start_row + index
        if any(verify_sheet.cell(row, column).value is None for column in (1, 5, 7)):
            errors.append(f"T{index + 1} 为空 (row {row})")
    verify.close()
    if errors:
        return False, "; ".join(errors)
    return True, f"OK (5 tiers, T1 sd={rows[0]['sd']}, T5 sd={rows[4]['sd']})"
