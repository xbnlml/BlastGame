#!/usr/bin/env python3
"""51-200 状态汇总工具

用法:
  python tools/stage_status.py              # 全部
  python tools/stage_status.py --simple      # 一行汇总
  python tools/stage_status.py 51,55,60      # 指定关卡
"""
import sys, os, json
from collections import defaultdict

STAGE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'stage-data')
TARGETS = r'C:\Users\Administrator\Documents\BlastGame\Assets\LvEditorConfig\lv_win_config_test.xlsx'
from openpyxl import load_workbook
import re


def get_difficulty_map():
    m = {}
    try:
        wb = load_workbook(TARGETS, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            try:
                lv = str(int(ws.cell(r, 1).value))
                diff = str(ws.cell(r, 9).value or '').lower()
                m[lv] = diff
            except Exception:
                pass
    except Exception:
        pass
    return m


def get_summary():
    sp = os.path.join(STAGE_DIR, '_summary.json')
    if os.path.exists(sp):
        try:
            return json.load(open(sp))
        except Exception:
            return {}
    return {}


def get_board_status():
    """从 board.md 读每关状态（真源）。progress.json 已废弃。
    ✅已入库=done / 🟡待调优|🔴需改关卡=need_tuning / 其他=no_data"""
    progress = {}
    board_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              'project-state', 'board.md')
    if os.path.exists(board_path):
        with open(board_path, encoding='utf-8') as f:
            for line in f:
                m = re.match(r'^\|\s*(\d{2,3})\s*\|\s*\w+\s*\|\s*([✅🟡🔴])[^\s|]*', line)
                if m:
                    lv = m.group(1)
                    status_mark = m.group(2)
                    if status_mark == '✅':
                        progress[lv] = {'status': 'done'}
                    elif status_mark in ('🟡', '🔴'):
                        progress[lv] = {'status': 'need_tuning'}
                    else:
                        progress.setdefault(lv, {'status': 'no_data'})
    return progress


def main():
    diffs = get_difficulty_map()
    summary = get_summary()

    progress = get_board_status()

    simple = '--simple' in sys.argv

    levels = [str(i) for i in range(51, 201)]
    if len(sys.argv) > 1 and not sys.argv[1].startswith('--'):
        spec = sys.argv[1]
        levels = []
        for part in spec.split(','):
            if '-' in part:
                a, b = part.split('-')
                levels.extend(str(i) for i in range(int(a), int(b) + 1))
            else:
                levels.append(part)

    done, need_tuning, no_data = [], [], []
    for lv in sorted(levels, key=int):
        st = progress.get(lv, {}).get('status')
        if st == 'done':
            done.append(lv)
        elif st == 'need_tuning':
            need_tuning.append(lv)
        else:
            no_data.append(lv)

    if simple:
        print(f'Done={len(done)}  Tuning={len(need_tuning)}  NoData={len(no_data)}  Total={len(levels)}')
        return

    print(f'51-200 状态汇总 ({len(levels)}关)')
    print('=' * 50)

    if done:
        print(f'\n完成 ({len(done)}):')
        for lv in done:
            d = diffs.get(lv, '?')
            s = summary.get('levels', {}).get(lv, {})
            bot_n = s.get("sources", {}).get("bot", 0)
            span_s = s.get("wr_span", "?")
            print("  %s (%s) bot=%s span=%s" % (lv.rjust(3), d.rjust(8), bot_n, span_s))
    print(f'\n待处理 ({len(need_tuning)}):')
    for lv in need_tuning:
        d = diffs.get(lv, '?')
        s = summary.get('levels', {}).get(lv, {})
        bot_n = s.get("sources", {}).get("bot", 0)
        span_s = s.get("wr_span", "?")
        print("  %s (%s) bot=%s span=%s" % (lv.rjust(3), d.rjust(8), bot_n, span_s))
    print(f'\n无数据 ({len(no_data)}):')
    for lv in no_data:
        d = diffs.get(lv, '?')
        print(f'  L{lv:>3s} ({d:>8})')


if __name__ == '__main__':
    main()