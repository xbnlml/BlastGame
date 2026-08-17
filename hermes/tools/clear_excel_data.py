#!/usr/bin/env python3
"""⑤ 清空 Excel 入库数据但保留行结构（关卡号/难度/档位 Tier1-Tier5）。

用途：关卡改配置后旧数据失效（2026-08-05 16 关案例），清数据列保留骨架。
只清数据列（4-9：胜率/sd/sc/ratios/of/备注），不动 1-3 列（关卡号/难度/档位）。

2026-08-05 审查修复：先验后写（结构验证通过才 save，失败不落盘）+ 备份名带日期。

用法：
  python tools/clear_excel_data.py --levels 102,110,119 --dry-run
  python tools/clear_excel_data.py --levels 102,110,119
  python tools/clear_excel_data.py --levels 101-125
"""
import argparse
import os
import shutil
import sys
from datetime import datetime

HERMES = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, HERMES)

XL_PATH = os.path.join(HERMES, '手动挑配置记录.xlsx')


def parse_levels(s):
    out = set()
    for part in str(s or '').split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            a, b = part.split('-')
            out.update(range(int(a), int(b) + 1))
        else:
            out.add(int(part))
    return sorted(out)


def _validate_structure(ws, lvs, dry_run):
    """先验：在内存 workbook 上验证 16 关 × 5 行结构（首行关卡号+难度、档位列）。
    通过返回 True，失败打印具体错误返回 False——真实模式在 save 前调用（先验后写）。
    """
    lv_rows = {}
    cur = None
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v is not None and isinstance(v, (int, float)):
            cur = int(v)
        if cur in lvs:
            tier = ws.cell(row, 3).value
            if tier and str(tier).startswith('Tier'):
                lv_rows.setdefault(cur, []).append(row)

    ok = True
    for lv in lvs:
        rows = lv_rows.get(lv, [])
        if len(rows) != 5:
            print(f'❌ L{lv}: {len(rows)} 行（应 5）')
            ok = False
            continue
        # 首行关卡号+难度（坑 119：只查 Tier1 行难度非空，Tier2-5 难度列本就空）
        if ws.cell(rows[0], 1).value != lv or ws.cell(rows[0], 2).value is None:
            print(f'❌ L{lv}: 首行关卡号/难度丢失')
            ok = False
            continue
        for i, r in enumerate(rows):
            if ws.cell(r, 3).value != f'Tier{i+1}':
                print(f'❌ L{lv}: 档位列错误 row{r}: {ws.cell(r, 3).value}')
                ok = False
    if ok:
        print(f'✅ 结构验证: {len(lvs)} 关 × 5 行保留'
              + ('' if dry_run else '，数据列已清空'))
    return ok


def clear_data(lvs, dry_run=False):
    import openpyxl

    if dry_run:
        print('DRY-RUN（不写文件，将清空以下行的数据列）')
    else:
        # 2026-08-05 审查修复：备份名带日期防跨天覆盖 + 语义标签
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        bak = XL_PATH.replace('.xlsx', f'_before_clear_{ts}.bak')
        shutil.copy2(XL_PATH, bak)
        print(f'备份: {os.path.basename(bak)}')

    wb = openpyxl.load_workbook(XL_PATH)
    ws = wb.active

    # 先验（P0：验证失败不落盘）
    if not _validate_structure(ws, lvs, dry_run):
        wb.close()
        if not dry_run:
            print('❌ 结构验证失败，未保存（Excel 未改动）')
        sys.exit(1)

    # 后写
    cleared = 0
    cur = None
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v is not None and isinstance(v, (int, float)):
            cur = int(v)
        if cur in lvs and ws.cell(row, 3).value and str(ws.cell(row, 3).value).startswith('Tier'):
            for c in range(4, 10):  # 数据列 4-9（胜率/sd/sc/ratios/of/备注）
                if not dry_run:
                    ws.cell(row, c).value = None
            cleared += 1
    print(f'{"将清空" if dry_run else "已清空"} {cleared} 行数据列')

    if not dry_run:
        wb.save(XL_PATH)
        print('Excel 已保存')
        # 写后回读验证
        wb2 = openpyxl.load_workbook(XL_PATH, read_only=True)
        ws2 = wb2.active
        lv_rows = {}
        cur2 = None
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and isinstance(row[0], (int, float)):
                cur2 = int(row[0])
            if cur2 in lvs and row[2] and str(row[2]).startswith('Tier'):
                lv_rows.setdefault(cur2, []).append(row)
        wb2.close()
        for lv in lvs:
            rows = lv_rows.get(lv, [])
            assert len(rows) == 5, f'L{lv} 回读行数错: {len(rows)}'
            for r in rows:
                assert all(x is None for x in r[3:9]), f'L{lv} 数据列未清空: {r[3:9]}'
        print('✅ 写后回读验证通过')
    wb.close()


def main():
    ap = argparse.ArgumentParser(description='清 Excel 数据列保留行结构（改关卡后数据失效）')
    ap.add_argument('--levels', required=True, help='关卡列表/区间')
    ap.add_argument('--dry-run', action='store_true', help='只检查不写')
    args = ap.parse_args()
    clear_data(parse_levels(args.levels), dry_run=args.dry_run)


if __name__ == '__main__':
    main()
