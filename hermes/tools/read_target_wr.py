#!/usr/bin/env python3
"""读 lv_win_config_test.xlsx 输出目标胜率表格

用法:
  python tools/read_target_wr.py              # 全部输出
  python tools/read_target_wr.py 51-100       # 指定范围
  python tools/read_target_wr.py 72,87,92     # 指定关卡
  python tools/read_target_wr.py 51-100 --json  # JSON 格式
"""

import sys, os, json
from openpyxl import load_workbook

REPO = os.environ.get('BLASTGAME_REPO', os.path.join(os.path.expanduser('~'), 'Documents', 'BlastGame'))
XLSX = os.path.join(REPO, 'Assets/LvEditorConfig/lv_win_config_test.xlsx')


def parse_levels(spec):
    levels = set()
    for part in spec.split(','):
        part = part.strip()
        if '-' in part:
            a, b = part.split('-')
            levels.update(range(int(a), int(b) + 1))
        else:
            levels.add(int(part))
    return sorted(levels)


def read_targets():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    targets = {}
    prev_lv = None
    for r in range(3, ws.max_row + 1):
        lv_raw = ws.cell(r, 1).value
        if lv_raw is not None:
            prev_lv = int(lv_raw)
        if prev_lv is None:
            continue
        t1 = ws.cell(r, 2).value
        if t1 is None:
            continue
        if prev_lv not in targets:
            targets[prev_lv] = {
                'diff': str(ws.cell(r, 9).value or ''),
                'tiers': [float(t1) * 100],
            }
            for c in range(3, 7):
                v = ws.cell(r, c).value
                targets[prev_lv]['tiers'].append(float(v) * 100 if v else 0)
    wb.close()
    return targets


def main():
    show_json = '--json' in sys.argv
    spec = None
    for arg in sys.argv[1:]:
        if arg != '--json':
            spec = arg
            break

    targets = read_targets()
    if spec:
        levels = parse_levels(spec)
    else:
        levels = sorted(targets.keys())

    if show_json:
        result = {}
        for lv in levels:
            if lv in targets:
                result[str(lv)] = targets[lv]
        print(json.dumps(result, indent=2, ensure_ascii=False))
        return

    print('  关    难度     T1    T2    T3    T4    T5')
    print('=' * 55)
    for lv in levels:
        if lv in targets:
            t = targets[lv]
            tiers = t['tiers']
            diff = t['diff']
            print(f'{lv:>4d} {diff:>12s} {tiers[0]:>5.0f}% {tiers[1]:>5.0f}% {tiers[2]:>5.0f}% {tiers[3]:>5.0f}% {tiers[4]:>5.0f}%')


if __name__ == '__main__':
    main()
